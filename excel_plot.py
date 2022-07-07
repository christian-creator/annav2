import pandas as pd
import numpy as np
import os
import sys
import datetime
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis
from boundaries import room_classes


def write_and_plot_excel_ANP0019(dataframe, room, slut_dato):
    if not os.path.exists('data/ANP0019/filtered'):
        os.makedirs('data/ANP0019/filtered')
    data = dataframe.to_numpy()
    # print(data)
    N, M = data.shape
    # Header for the numpy array
    header = np.array(["date"] + ["Point {}".format(point+1)
                                  for point in range(M-1)] + ["Alert CFU/m3 {}".format(room_classes[room]["ANP0019_alert"])]).reshape(1, -1)
    # Rearranging data due to the dates not being indexes instead of actual values
    # print(N)
    
    dates = dataframe.index.values.reshape(N, -1)
    data = np.concatenate((dates, data), axis=1)
    # Adding the header
    data = np.concatenate((header, data), axis=0)
    # Addding the end date
    end_date = [int(x) for x in slut_dato.split("-")]
    end_date = datetime.date(*end_date)
    last_element = np.array([end_date] + ["" for x in range(len(data[0])-2)] + [data[1][-1]]).reshape(1,len(data[0]))   
    data = np.concatenate((data, last_element), axis=0)
 
    
    # Shape of the final data structure
    N, M = data.shape
    # Convert to list
    rows = data.tolist()
    # Write to excel
    wb = Workbook()
    ws = wb.active
    cs = wb.create_chartsheet()
    for row in rows:
        ws.append(row)
    data = Reference(ws, min_col=2, min_row=1, max_col=M, max_row=N)
    # Chart with date axis
    c1 = LineChart()
    c1.title = "Class {} - Room {} \n Microbial air sampling".format(room_classes[room]["class"],room.replace("_","."))
    # c1.title = "Microbial air sampling: {} class {}".format(
    #     room, room_classes[room]["class"])
    c1.style = 12
    c1.y_axis.title = "CFU"
    c1.y_axis.crossAx = 500
    c1.x_axis = DateAxis(crossAx=100)
    c1.x_axis.number_format = 'yyyy-mm-dd'
    c1.x_axis.majorTimeUnit = "days"
    # Units between the ticks
    c1.x_axis.majorUnit = 8
    c1.x_axis.title = "Date"
    c1.add_data(data, titles_from_data=True)
    dates = Reference(ws, min_col=1, min_row=2, max_row=N)
    colors = ["0FDC30", "4F81BD", "FF3355",
              "4CDFD3", "FF33D9", "FF3F33", "7030A0"]
    alert_color = "F79646"
    for i in range(len(rows[1])-1):
        if i == len(rows[1])-2:
            s = c1.series[i]
            s.smooth = True
            s.graphicalProperties.line.width = 40000  # width in EMUs
            s.graphicalProperties.line.solidFill = alert_color
        else:
            s = c1.series[i]
            s.marker.symbol = "diamond"
            s.marker.graphicalProperties.solidFill = colors[i]
            s.marker.graphicalProperties.line.solidFill = colors[i]
            s.graphicalProperties.line.width = 40000  # width in EMUs
            s.graphicalProperties.line.solidFill = colors[i]
    c1.y_axis.scaling.min = 0
    c1.y_axis.scaling.max = room_classes[room]["ANP0019_action"]
    c1.set_categories(dates)
    cs.add_chart(c1)
    wb.save("data/ANP0019/filtered/{}.xlsx".format(room))


def write_and_plot_excel_ANP0020(dataframe, room, slut_dato):
    if not os.path.exists('data/ANP0020/filtered'):
        os.makedirs('data/ANP0020/filtered')
    # Converting data to correct dataframe with header and dates
    data = dataframe.to_numpy()
    N, M = data.shape
    N_points = int((M-2)/2)
    # Header for the numpy array
    header = np.array(["date"] + ["0.5 \u03BCm \nPoint {}".format(point+1) for point in range(N_points)] +
                      ["5.0 \u03BCm \nPoint {}".format(point+1) for point in range(N_points)] + ["alert 0.5 \n{}".format(room_classes[room]["ANP0020_alert"][0]), "alert 5.0 \n{}".format(room_classes[room]["ANP0020_alert"][1])]).reshape(1, -1)
    # Rearranging data due to the dates not being indexes instead of actual values
    dates = dataframe.index.values.reshape(N, -1)
    data = np.concatenate((dates, data), axis=1)
    # Adding the header
    data = np.concatenate((header, data), axis=0)
    # Adding the last date
    end_date = [int(x) for x in slut_dato.split("-")]
    end_date = datetime.date(*end_date)
    last_element = np.array([end_date] + ["" for x in range(len(data[0])-3)] + [data[1][-2]] + [data[1][-1]]).reshape(1,len(data[0]))   
    data = np.concatenate((data, last_element), axis=0)

    # Shape of the final data structure
    N, M = data.shape
    # Convert to list
    rows = data.tolist()
    # Write to excel
    wb = Workbook()
    ws = wb.active
    cs = wb.create_chartsheet()
    for row in rows:
        ws.append(row)
    # Adding the data from the different sections of the data
    # Dates
    dates = Reference(ws, min_col=1, min_row=2, max_row=N)
    # Alerts
    # 0.5
    alert_0_5 = Reference(ws,min_col=M-1, min_row=1, max_row=N)
    alert_0_5_color = "F79646"
    # 5.0
    alert_5_0 = Reference(ws,min_col=M, min_row=1, max_row=N)
    alert_5_0_color = "4F81BD"
    # Adding the actual data
    # C1 is 5.0
    c1 = LineChart()
    data = Reference(ws, min_col= 2 + N_points, min_row= 1, max_col= (2 + N_points) + (N_points-1), max_row=N)
    c1.style = 12
    c1.y_axis.title = "Counts, 5.0 \u03BCm"
    c1.y_axis.majorGridlines = None
    c1.y_axis.crossAx = 500
    c1.y_axis.scaling.min = 0
    c1.y_axis.scaling.max = room_classes[room]["ANP0020_action"][1]
    c1.x_axis = DateAxis(crossAx=100)
    c1.x_axis.number_format = 'yyyy-mm-dd'
    c1.x_axis.majorTimeUnit = "days"
    # Units between the ticks
    c1.x_axis.majorUnit = 8
    c1.x_axis.title = "Date"
    c1.add_data(data, titles_from_data=True)
    c1.add_data(alert_5_0, titles_from_data=True)
    c1.set_categories(dates)
    c1.title = "Class {} - Room {} \n Particle counting".format(room_classes[room]["class"],room.replace("_","."))
    # c1.title = "Particle counting {} Class {}".format(room,room_classes[room]["class"])
    colors = ["0FDC30", "FF3355", "4CDFD3", "FF33D9", "FF3F33", "7030A0","#A9DFBF"]
    # Color the series
    for i in range(len(c1.series)):
        if i < (len(c1.series) - 1):
            s = c1.series[i]
            s.marker.symbol = "diamond"
            s.marker.graphicalProperties.solidFill = colors[i]
            s.marker.graphicalProperties.line.solidFill = colors[i]
            s.graphicalProperties.line.width = 40000  # width in EMUs
            s.graphicalProperties.line.solidFill = colors[i]
        else:
            s = c1.series[i]
            s.smooth = True
            s.graphicalProperties.line.width = 40000  # width in EMUs
            s.graphicalProperties.line.solidFill = alert_5_0_color

    # C2 is 0.5
    c2 = LineChart()
    data = Reference(ws, min_col=2, min_row=1, max_col=2+(N_points-1), max_row=N)
    c2.style = 12
    c2.y_axis.title = "Counts, 0.5 \u03BCm"
    c2.y_axis.crossAx = 500
    c2.y_axis.scaling.min = 0
    c2.y_axis.scaling.max = room_classes[room]["ANP0020_action"][0]
    c2.y_axis.axId = 200
    c2.x_axis = DateAxis(crossAx=100)
    c2.x_axis.number_format = 'yyyy-mm-dd'
    c2.x_axis.majorTimeUnit = "days"
    # Units between the ticks
    c1.x_axis.majorUnit = 8
    c2.add_data(data, titles_from_data=True)
    c2.add_data(alert_0_5, titles_from_data=True)
    c2.set_categories(dates)
    # Color the different series
    colors = ["008000","A93226","7D3C98","1ABC9C","EC7063","BB8FCE","ABEBC6"]
    for i in range(len(c2.series)):
            if i < (len(c2.series) - 1):
                s = c2.series[i]
                s.marker.symbol = "diamond"
                s.marker.graphicalProperties.solidFill = colors[i]
                s.marker.graphicalProperties.line.solidFill = colors[i]
                s.graphicalProperties.line.width = 40000  # width in EMUs
                s.graphicalProperties.line.solidFill = colors[i]
            else:
                s = c2.series[i]
                s.smooth = True
                s.graphicalProperties.line.width = 40000  # width in EMUs
                s.graphicalProperties.line.solidFill = alert_0_5_color

    # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
    c1.y_axis.crosses = "max"
    c1 += c2

    # ws.add_chart(c1, "D4")
    cs.add_chart(c1)
    wb.save("data/ANP0020/filtered/{}.xlsx".format(room))
    # wb.save("secondary.xlsx")
    # sys.exit(1)





def write_and_plot_excel_ANP0230(dataframe, room, slut_dato):
    if not os.path.exists('data/ANP0230/filtered'):
        os.makedirs('data/ANP0230/filtered')
    data = dataframe.to_numpy()
    # print(data)
    N, M = data.shape
    # Header for the numpy array
    header = np.array(["date"] + ["Point {}".format(point+1)
                                  for point in range(M-1)] + ["Alert CFU/m3 {}".format(room_classes[room]["ANP0230_alert"])]).reshape(1, -1)
    # Rearranging data due to the dates not being indexes instead of actual values
    # print(N)
    
    dates = dataframe.index.values.reshape(N, -1)
    data = np.concatenate((dates, data), axis=1)
    # Adding the header
    data = np.concatenate((header, data), axis=0)
    # Addding the end date
    end_date = [int(x) for x in slut_dato.split("-")]
    end_date = datetime.date(*end_date)
    last_element = np.array([end_date] + ["" for x in range(len(data[0])-2)] + [data[1][-1]]).reshape(1,len(data[0]))   
    data = np.concatenate((data, last_element), axis=0)
 
    
    # Shape of the final data structure
    N, M = data.shape
    # Convert to list
    rows = data.tolist()
    # Write to excel
    wb = Workbook()
    ws = wb.active
    cs = wb.create_chartsheet()
    for row in rows:
        ws.append(row)
    data = Reference(ws, min_col=2, min_row=1, max_col=M, max_row=N)
    # Chart with date axis
    c1 = LineChart()
    c1.title = "Class {} - Room {} \n Contact Plates".format(room_classes[room]["class"],room.replace("_","."))
    # c1.title = "Microbial air sampling: {} class {}".format(
    #     room, room_classes[room]["class"])
    c1.style = 12
    c1.y_axis.title = "CFU"
    c1.y_axis.crossAx = 500
    c1.x_axis = DateAxis(crossAx=100)
    c1.x_axis.number_format = 'yyyy-mm-dd'
    c1.x_axis.majorTimeUnit = "days"
    # Units between the ticks
    c1.x_axis.majorUnit = 8
    c1.x_axis.title = "Date"
    c1.add_data(data, titles_from_data=True)
    dates = Reference(ws, min_col=1, min_row=2, max_row=N)
    colors = ["0FDC30", "4F81BD", "FF3355",
              "4CDFD3", "FF33D9", "FF3F33", "7030A0"]
    alert_color = "F79646"
    for i in range(len(rows[1])-1):
        if i == len(rows[1])-2:
            s = c1.series[i]
            s.smooth = True
            s.graphicalProperties.line.width = 40000  # width in EMUs
            s.graphicalProperties.line.solidFill = alert_color
        else:
            s = c1.series[i]
            s.marker.symbol = "diamond"
            s.marker.graphicalProperties.solidFill = colors[i]
            s.marker.graphicalProperties.line.solidFill = colors[i]
            s.graphicalProperties.line.width = 40000  # width in EMUs
            s.graphicalProperties.line.solidFill = colors[i]
    c1.y_axis.scaling.min = 0
    c1.y_axis.scaling.max = room_classes[room]["ANP0230_action"]
    c1.set_categories(dates)
    cs.add_chart(c1)
    wb.save("data/ANP0230/filtered/{}.xlsx".format(room))